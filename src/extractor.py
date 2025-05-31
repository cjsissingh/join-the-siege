"""
This module provides functions for extracting text from various file formats.
"""
from io import BytesIO
import logging
from werkzeug.datastructures import FileStorage

# Import libraries for content extraction
import magic
from PyPDF2 import PdfReader
from PIL import Image
import pytesseract
import docx
import openpyxl

# Import configuration
from src.config import LOG_LEVEL

logger = logging.getLogger(__name__)
logging.basicConfig(
    level=LOG_LEVEL,
)

def extract_text_from_file(file: FileStorage) -> str | None:
    """
    Extracts text from various file formats.

    Note:
    - For image OCR (PNG, JPG, TIFF), Tesseract OCR must be installed on the system
    and accessible in the system's PATH.
    - `python-magic` relies on `libmagic` which might need separate installation
    depending on the OS.

    Args:
        file: A FileStorage object representing the uploaded file.

    Returns:
        A string containing the extracted text, or None if extraction fails
        or the file type is unsupported.
    """
    try:
        file.seek(0)
        file_bytes = file.read()
        # Reset stream pointer for subsequent parsers
        file.seek(0)

        mime_type = magic.from_buffer(file_bytes, mime=True)
        logger.debug("Detected MIME type: %s for file %s", mime_type, file.filename)

        text_parts = []

        try:
            # Extract text based on MIME type
            match mime_type:
                # PDF
                case "application/pdf":
                    reader = PdfReader(file)
                    for page in reader.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text_parts.append(page_text)
                    logger.debug("Successfully extracted text from PDF: %s", file.filename)

                # Images
                case "image/png" | "image/jpeg" | "image/tiff":
                    img = Image.open(BytesIO(file_bytes))
                    extracted_text = pytesseract.image_to_string(img)
                    if extracted_text:
                        text_parts.append(extracted_text)
                    logger.debug("Successfully extracted text from image (OCR): %s", file.filename)

                # DOCX
                case "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                    document = docx.Document(file)
                    for para in document.paragraphs:
                        if para.text:
                            text_parts.append(para.text)
                    logger.debug("Successfully extracted text from DOCX: %s", file.filename)

                # XLS
                case "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                    workbook = openpyxl.load_workbook(file)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value and isinstance(cell.value, str):
                                    text_parts.append(cell.value)
                    logger.debug("Successfully extracted text from XLSX: %s", file.filename)

                # Plain text
                case "text/plain":
                    # file_bytes were read earlier, decode them
                    text_content = file_bytes.decode('utf-8', errors='ignore')
                    text_parts.append(text_content)
                    logger.debug(
                        "Successfully extracted text from plain text file: %s", 
                        file.filename
                    )

                 # Unsupported
                case _:
                    logger.warning(
                        "Unsupported MIME type for text extraction: %s for file %s",
                        mime_type,
                        file.filename,
                    )
                    return None

        # Catch a generic exception from the file parsers
        except Exception as e:
            logger.error(
                lambda: f"Error extracting text from {file.mimetype} for {file.filename}: {e}", exc_info=True
            )
            return None

        return "\n".join(text_parts).strip() if text_parts else None

    except Exception as e:
        logger.error(
            lambda: f"Generic error in extract_text_from_file for {file.filename}: {e}", exc_info=True
        )
        return None

